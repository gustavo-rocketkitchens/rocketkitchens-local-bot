from request_app import GetMenuItem
import menu_category
import logging
from playwright.sync_api import sync_playwright
import time
import csv
import re
import pandas as pd
import openpyxl


logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s:%(levelname)s:%(message)s')
file_handler = logging.FileHandler('app.log')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)


class MarketingAnalysis(GetMenuItem):

    def __init__(self):
        super().__init__()
        self.menu = None
        self.dict_area = None

    # Get the URL of the Area
    def input_area(self, area):
        logger.info(f"Getting the URL for {area}")
        self.menu = GetMenuItem()
        self.dict_area = self.menu.urls_per_area
        url = self.dict_area[area]
        url = 'https://www.talabat.com/' + url
        logger.info(f"The URL for {area} is: {url}")
        return url

    # Get the restaurants Info by Cuisine
    # Details for restaurant
    # Title
    # Ratings and new section
    def input_cuisine(self, cuisine, url):
        logger.info(f"Searching for {cuisine} in {url}")
        self.menu = GetMenuItem()
        self.all_cuisine = self.menu.all_cuisine
        restaurants = {}

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            page.goto(url)
            time.sleep(.5)
            logger.info(f"Opened the browser and navigated to {url}")
            search_input = page.query_selector("//input[@placeholder='Search Restaurants']")
            search_input.fill(cuisine)
            time.sleep(.2)  # Wait for 2 seconds to ensure page is loaded
            logger.info(f"Filled in the cuisine search input with {cuisine}")


            restaurant_list = page.query_selector_all("//div[@class='vendor-card']")
            for index, restaurant in enumerate(restaurant_list):
                restaurant_title = restaurant.query_selector(".content")
                # cuisines_section = restaurant.query_selector(".cuisines-section.pb-1.truncate")
                # ratings_section = restaurant.query_selector(".ratings-and-new-section.pb-1.d-flex")
                # info_section = restaurant.query_selector(".info-section.pb-1.f-14.delivery-info.d-flex")
                # tlb_badge_section = restaurant.query_selector(
                #     ".tlb-badge-section.d-flex.align-items-center.flex-wrap")
                logger.info(f"Details for restaurant {index + 1}:")
                logger.info(f"Title: {restaurant_title.inner_text()}")
                # logger.info(f"Cuisine: {cuisines_section.inner_text()}")
                # logger.info(f"Ratings and new section: {ratings_section.inner_text()}")
                # logger.info(f"Info section: {info_section.inner_text()}")
                # logger.info(f"TLB badge section: {tlb_badge_section.inner_text()}")
                restaurant_details = {
                    "Title": restaurant_title.inner_text(),
                    # "Cuisine": cuisines_section.inner_text(),
                    # "Ratings and new section": ratings_section.inner_text(),
                    # "Info section": info_section.inner_text(),
                    # "TLB badge section": tlb_badge_section.inner_text()
                }
                restaurants[restaurant_title.inner_text()] = restaurant_details
                time.sleep(2)

        return restaurants

    # Get the restaurants addresses by Cuisine
    # Title
    # URL
    def output_restaurants_url(self, cuisine, url):
        logger.info(f"Searching for restaurants URL'S in {url}")
        self.menu = GetMenuItem()
        self.all_cuisine = self.menu.all_cuisine
        restaurants = {}

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            page = browser.new_page()
            page.goto(url)
            time.sleep(.2)
            logger.info(f"Opened the browser and navigated to {url}")
            search_input = page.query_selector("//input[@placeholder='Search Restaurants']")
            search_input.fill(cuisine)
            time.sleep(2)  # Wait for 2 seconds to ensure page is loaded
            logger.info(f"Filled in the cuisine search input with {cuisine}")

            restaurant_urls = page.query_selector_all("//a[@data-test='restaurant-a']")
            if restaurant_urls:
                for restaurant_url in restaurant_urls:
                    restaurant_title = restaurant_url.query_selector(".content h2")
                    url = restaurant_url.get_attribute("href")
                    restaurant_details = {
                        "Title": restaurant_title.inner_text(),
                        "URL": url
                    }
                    logger.info(f"restaurant_details: {url}")
                    logger.info(f"Title: {restaurant_title}")
                    restaurants[restaurant_title.inner_text()] = restaurant_details

        # logger.info(f"List of Restaurants URL's: {restaurants}")

        return restaurants

    # def get_menu_categories(self, restaurants, cuisine, quantity=None, output_file='menu.txt'):
    #     self.menu_items = {}
    #     self.result_dict = {}
    #     self.menu_list = []
    #
    #     num = 0
    #
    #     for restaurant in restaurants.values():
    #         # Title: Url
    #         logger.info(f"Entering {restaurant}")
    #         url = "https://www.talabat.com" + restaurant['URL']
    #         self.menu_items = menu_category.scrape_menu_items(url, cuisine)
    #         time.sleep(0.1)
    #         logger.info(f"Menu Items Extracted for {cuisine} in {restaurant['URL']}")
    #         num += 1
    #         if quantity and num == quantity:
    #             break
    #         logger.info(f"Apendding {num}: Menu Items = {self.menu_items}")
    #         self.menu_list.append(restaurant)
    #         self.menu_list.append(self.menu_items)
    #
    #     print(f"List of dicts is {self.menu_list}")
    #
    #     # Write to file
    #     with open(output_file, 'w', encoding='utf-8') as file:
    #         for item in self.menu_list:
    #             for key, value in item.items():
    #                 file.write(f"{key}: {value}\n")
    #             file.write('\n')
    #
    #     logger.info(f"List of dicts written to {output_file}")


    def get_menu_categories(self, restaurants, cuisine, quantity=None, output_file='menu.xlsx'):
        menu_list = []

        num = 0

        for restaurant in restaurants.values():
            if quantity and num == quantity:
                break

            url = "https://www.talabat.com" + restaurant['URL']
            menu_items = menu_category.scrape_menu_items(url, cuisine)
            menu_dict = {'Title': restaurant['Title'], 'URL': url}

            for key, value in menu_items.items():
                menu_item_dict = {}

                for item in value:
                    menu_item_dict['name'] = item['name']
                    menu_item_dict['description'] = item['description']
                    menu_item_dict['price'] = item['price']
                    menu_dict[key] = [menu_item_dict]

            menu_list.append(menu_dict)
            num += 1

        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write header row
        ws.append(['Restaurant Title', 'URL', 'Food Name', 'Description', 'Price'])

        # Iterate over the menu items and write them to the worksheet
        row_num = 2
        for item in menu_list:
            # Write the restaurant title and URL to the worksheet
            ws.cell(row=row_num, column=1, value=item['Title'])
            ws.cell(row=row_num, column=2, value=item['URL'])

            # Iterate over the menu items and write them to the worksheet
            for key, value in item.items():
                if key not in ['Title', 'URL']:
                    for menu_item in value:
                        ws.cell(row=row_num, column=3, value=menu_item['name'])
                        ws.cell(row=row_num, column=4, value=menu_item['description'])
                        ws.cell(row=row_num, column=5, value=menu_item['price'])
                        row_num += 1

        output_file = 'menu_' + cuisine + '_in_' + area + '.xlsx'
        # Save the workbook
        wb.save(output_file)


if __name__ == '__main__':
    mkt = MarketingAnalysis()
    # area = 'Business Bay'
    area = 'Dubai Motor City'
    cuisine = 'Sushi'
    # logger.info(f"Getting details for restaurants in {area} serving {cuisine} cuisine")
    url = mkt.input_area(area)

    mkt.input_cuisine(cuisine, url)
    logger.info(f"Finished retrieving restaurant details for {cuisine} cuisine in {url}")
    # time.sleep(10)
    restaurants = mkt.output_restaurants_url(cuisine, url)
    logger.info(f"Finished retrieving restaurant URL's for {cuisine}")
    # #
    # logger.info(f"Getting Menu Category Info for restaurants in {area} serving {cuisine} cuisine")
    mkt.get_menu_categories(restaurants=restaurants, cuisine=cuisine, quantity=1)
    time.sleep(.2)

    # logger.info(f"Getting and Saving Menu Category Info for restaurants in {area} serving {cuisine} cuisine")
    # time.sleep(2)

